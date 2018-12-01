import requests
import json
import xlsxwriter
import openpyxl 
from datetime import datetime, timedelta

#function to write json object file
def writecsvdata(data):
	year = data['items'][0]['year']
	month = data['items'][0]['month']
	day = data['items'][0]['day']

	
	#wb = openpyxl.load_workbook('wikipediadata.xlsx')
	#sheet = wb.active
	

	
	for article in data['items'][0]['articles']:
		temp = [article['article'], article['views'], day, month, year]
		if article['article'] == "Main_Page" or int(article['views']) < 50000:
			pass
		else:
			sheet.append(temp)	
		

	print(sheet.max_row)
	#wb.save('wikipediadata.xlsx')

# First write function

def firstwrite(data):
	#print(data)
	year = data['items'][0]['year']
	month = data['items'][0]['month']
	day = data['items'][0]['day']

	
	workbook = xlsxwriter.Workbook('wikipediadata.xlsx')
	csvwriter = workbook.add_worksheet()

	row = 1
	col = 0
	for article in data['items'][0]['articles']:
		csvwriter.write(row, col, article['article'])
		csvwriter.write(row, col + 1, article['views'])
		csvwriter.write(row, col + 2, day)
		csvwriter.write(row, col + 3, month)
		csvwriter.write(row, col + 4, year)

		row += 1

	workbook.close()

	




# create a list with all dates to processed
start_date=datetime( 2015, 10, 17 )
end_date=datetime( 2018, 10, 17 )
d=start_date
dates=[ start_date ]
while d < end_date:
	d += timedelta(days=1)
	dates.append(str(datetime.strftime(d,'%x')))
	
print(len(dates))
#print(dates[1].split('/'))

wb = openpyxl.load_workbook('wikipediadata.xlsx')
sheet = wb.active
for i in range(1,1097):
	print("i: ", i)
	nd = dates[i].split('/')
	mm = nd[0]
	dd = nd[1]
	yy = "20" + nd[2]
	#print(yy + '/' + mm + '/' + dd)

	#URL = "https://wikimedia.org/api/rest_v1/metrics/pageviews/top/en.wikipedia.org/all-access/2016/01/01"
	URL = "https://wikimedia.org/api/rest_v1/metrics/pageviews/top/en.wikipedia.org/all-access/" + yy + '/' + mm + '/' + dd
	r = requests.get(url = URL)
	data = r.json()
	writecsvdata(data)
wb.save('wikipediadata.xlsx')

# URL = "https://wikimedia.org/api/rest_v1/metrics/pageviews/top/en.wikipedia.org/all-access/2016/01/01"
# r = requests.get(url = URL)
# data = r.json()
# firstwrite(data)

